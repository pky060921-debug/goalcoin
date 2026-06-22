from flask import Blueprint, request, jsonify, send_file
import sqlite3
import threading
import uuid
import json
import logging
import traceback
import os
import sys
import random
import re
import time
import requests
import subprocess
import ast
from datetime import datetime

from google import genai
from google.genai import types

from config import GEMINI_API_KEYS, TASK_STATUS
from database import get_db_connection
from services.parser import parse_html_3col_law, normalize_text, clean_korean_law_text, get_next_review_time

try:
    import fitz  # PyMuPDF
except ImportError:
    logging.error("PyMuPDF(fitz) 라이브러리가 설치되지 않았습니다.")

api_bp = Blueprint('api', __name__)

# ==========================================
# 💡 GOAL 코인 발행 (Sui Blockchain 연동)
# ==========================================
SUI_PACKAGE_ID = "YOUR_PACKAGE_ID_HERE"
SUI_TREASURY_CAP_ID = "YOUR_TREASURY_CAP_ID_HERE"

def mint_goal_coin_to_user(wallet_address, amount=10):
    if SUI_PACKAGE_ID == "YOUR_PACKAGE_ID_HERE":
        logging.warning("⚠️ 스마트 컨트랙트 ID 미설정: Goalcoin 지급이 생략됩니다.")
        return False
        
    try:
        raw_amount = str(amount * 1000000000) # Decimals 9
        cmd = [
            "sui", "client", "call", 
            "--package", SUI_PACKAGE_ID,
            "--module", "goal", 
            "--function", "mint_reward",
            "--args", SUI_TREASURY_CAP_ID, raw_amount, wallet_address,
            "--gas-budget", "50000000",
            "--json"
        ]
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode == 0:
            logging.info(f"🪙 {wallet_address} 에게 {amount} GOAL 지급 완료!")
            return True
        else:
            logging.error(f"❌ GOAL 지급 실패: {result.stderr}")
            return False
    except Exception as e:
        logging.error(f"❌ 코인 발행 에러: {str(e)}")
        return False

# ==========================================
# 💡 로컬 AI (Ollama) 고도화 엔진
# ==========================================
current_api_key_index = 0

def generate_ollama_json(prompt, model="gemma4:26b", temperature=0.1):
    try:
        url = "http://localhost:11434/api/generate"
        payload = {
            "model": model,
            "prompt": prompt,
            "stream": False,
            "format": "json",
            "options": {
                "temperature": temperature,
                "num_ctx": 16384,
                "num_predict": 1024,
                "top_k": 40,
                "top_p": 0.9,
                "repeat_penalty": 1.15
            }
        }
        response = requests.post(url, json=payload, timeout=300)
        response.raise_for_status()
        return response.json().get("response", "{}")
    except Exception as e:
        print(f"\n[🔥 로컬 AI (Ollama) 통신 에러]\n{e}\n", file=sys.stderr, flush=True)
        raise e

def generate_ollama_text(prompt, model="gemma4:26b", temperature=0.1):
    """채팅 모델용 /api/chat 엔드포인트 사용."""
    try:
        url = "http://localhost:11434/api/chat"
        payload = {
            "model": model,
            "messages": [{"role": "user", "content": prompt}],
            "stream": False,
            "options": {
                "temperature": temperature,
                "num_ctx": 16384,
                "num_predict": 1024,
                "top_k": 40,
                "top_p": 0.9,
                "repeat_penalty": 1.15
            }
        }
        response = requests.post(url, json=payload, timeout=300)
        response.raise_for_status()
        return response.json().get("message", {}).get("content", "")
    except Exception as e:
        print(f"\n[🔥 Ollama chat 통신 에러]\n{e}\n", file=sys.stderr, flush=True)
        raise e

def sanitize_json_string_values(text):
    """JSON 문자열 값 안의 raw 개행/탭을 이스케이프 처리."""
    result = []
    in_string = False
    escape_next = False
    for ch in text:
        if escape_next:
            result.append(ch)
            escape_next = False
            continue
        if ch == '\\':
            result.append(ch)
            escape_next = True
            continue
        if ch == '"':
            in_string = not in_string
            result.append(ch)
            continue
        if in_string:
            if ch == '\n':
                result.append('\\n')
            elif ch == '\r':
                result.append('\\r')
            elif ch == '\t':
                result.append('\\t')
            else:
                result.append(ch)
        else:
            result.append(ch)
    return ''.join(result)

def clean_and_parse_json(response_text):
    try:
        text = response_text.strip()
        text = re.sub(r'```json\s*', '', text)
        text = re.sub(r'```\s*', '', text).strip()

        start_idx_arr = text.find('[')
        end_idx_arr = text.rfind(']')
        start_idx_obj = text.find('{')
        end_idx_obj = text.rfind('}')

        if start_idx_arr != -1 and end_idx_arr != -1 and (start_idx_obj == -1 or start_idx_arr < start_idx_obj):
            clean_text = text[start_idx_arr:end_idx_arr+1]
        elif start_idx_obj != -1 and end_idx_obj != -1:
            clean_text = text[start_idx_obj:end_idx_obj+1]
        else:
            clean_text = text

        try:
            return json.loads(clean_text)
        except json.decoder.JSONDecodeError:
            pass

        try:
            sanitized = sanitize_json_string_values(clean_text)
            return json.loads(sanitized)
        except json.decoder.JSONDecodeError:
            pass

        try:
            fixed = re.sub(r"([{,]\s*)'([^']+)'(\s*:)", r'\1"\2"\3', clean_text)
            sanitized = sanitize_json_string_values(fixed)
            return json.loads(sanitized)
        except Exception:
            pass

        try:
            python_dict = ast.literal_eval(clean_text)
            return json.loads(json.dumps(python_dict, ensure_ascii=False))
        except Exception:
            pass

        raise ValueError(f"JSON 파싱 실패. Raw: {response_text[:200]}")
    except Exception as e:
        print(f"JSON Parsing Error. Raw Text: {response_text[:300]}", file=sys.stderr)
        raise e

# ==========================================
# 💡 DB 초기화 (user_settings 포함)
# ==========================================
def init_golden_db():
    try:
        conn = get_db_connection()
        conn.execute('''CREATE TABLE IF NOT EXISTS golden_exams (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            wallet_address TEXT,
            title TEXT,
            question TEXT,
            options_json TEXT,
            answer TEXT,
            explanation TEXT,
            category TEXT
        )''')
        conn.execute('''CREATE TABLE IF NOT EXISTS pending_exams (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            wallet_address TEXT,
            filename TEXT,
            chunks_json TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        conn.execute('''CREATE TABLE IF NOT EXISTS user_settings (
            wallet_address TEXT PRIMARY KEY,
            custom_stopwords TEXT
        )''')
        
        try: conn.execute('ALTER TABLE categories ADD COLUMN sort_order INTEGER DEFAULT 999999')
        except: pass
        try: conn.execute('ALTER TABLE cards ADD COLUMN sort_order INTEGER DEFAULT 999999')
        except: pass
            
        try: conn.execute('ALTER TABLE golden_exams ADD COLUMN search_process TEXT DEFAULT ""')
        except: pass
        try: conn.execute('ALTER TABLE golden_exams ADD COLUMN referenced_laws TEXT DEFAULT ""')
        except: pass
        try: conn.execute('ALTER TABLE categories ADD COLUMN folder_name TEXT DEFAULT "기본 폴더"')
        except: pass
        try: conn.execute('ALTER TABLE cards ADD COLUMN folder_name TEXT DEFAULT "기본 폴더"')
        except: pass
        try: conn.execute('ALTER TABLE cards ADD COLUMN memo TEXT DEFAULT ""')
        except: pass
        try: conn.execute('ALTER TABLE cards ADD COLUMN best_time REAL DEFAULT 999.0')
        except: pass
        try: conn.execute('ALTER TABLE user_settings ADD COLUMN ai_rules TEXT DEFAULT ""')
        except: pass
        try: conn.execute('ALTER TABLE pending_exams ADD COLUMN answers_json TEXT DEFAULT "[]"')
        except: pass
        try: conn.execute('ALTER TABLE user_settings ADD COLUMN custom_abbrs TEXT DEFAULT "{}"')
        except: pass
        try: conn.execute('ALTER TABLE user_settings ADD COLUMN custom_inclusions TEXT DEFAULT "[]"')
        except: pass
            
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"\n[🔥 DB 초기화 에러]\n{traceback.format_exc()}\n", file=sys.stderr, flush=True)

init_golden_db()

@api_bp.route('/health', methods=['GET', 'OPTIONS'])
def health_check():
    return jsonify({"status": "alive"}), 200

@api_bp.route('/task-status')
def task_status():
    task_id = request.args.get('task_id')
    if task_id in TASK_STATUS:
        return jsonify(TASK_STATUS[task_id])
    return jsonify({"status": "not_found"}), 404

# ==========================================
# 💡 폴더 및 파일 관리 라우터
# ==========================================
@api_bp.route('/delete-law-file', methods=['POST'])
def delete_law_file():
    try:
        data = request.json or {}
        folder_name = data.get('folder_name')
        wallet_address = data.get('wallet_address')
        if not folder_name or not wallet_address: return jsonify({"error": "삭제 권한이 없습니다."}), 400

        conn = get_db_connection()
        conn.execute("DELETE FROM categories WHERE folder_name = ? AND wallet_address = ?", (folder_name, wallet_address))
        conn.commit()
        conn.close()
        return jsonify({"message": "법령 파일 삭제 완료"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@api_bp.route('/delete-folder', methods=['POST'])
def delete_folder():
    return delete_law_file()

@api_bp.route('/rename-folder', methods=['POST'])
def rename_folder():
    try:
        data = request.json
        wallet_address = data.get('wallet_address')
        old_folder_name = data.get('old_folder_name')
        new_folder_name = data.get('new_folder_name')
        if not wallet_address or not old_folder_name: return jsonify({"error": "정보 누락"}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE categories SET folder_name = ? WHERE wallet_address = ? AND folder_name = ?", (new_folder_name, wallet_address, old_folder_name))
        conn.commit()
        conn.close()
        return jsonify({"message": "폴더명 변경 완료"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@api_bp.route('/update-category-folder', methods=['POST'])
def update_category_folder():
    try:
        data = request.json
        wallet_address = data.get('wallet_address')
        cat_id = data.get('id')
        new_folder_name = data.get('new_folder_name')
        if not wallet_address or not cat_id: return jsonify({"error": "정보 누락"}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE categories SET folder_name = ? WHERE id = ? AND wallet_address = ?", (new_folder_name, cat_id, wallet_address))
        conn.commit()
        conn.close()
        return jsonify({"message": "폴더 이동 완료"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@api_bp.route('/update-category-text', methods=['POST'])
def update_category_text():
    try:
        data = request.json or {}
        wallet_address = data.get('wallet_address')
        cat_id = data.get('id')
        content = data.get('content') # 필드명을 DB 컬럼과 맞춰 content로 수신
        
        if not wallet_address or not cat_id:
            return jsonify({"error": "필수 정보가 누락되었습니다."}), 400

        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 데이터베이스의 categories 테이블 내 content 컬럼을 업데이트합니다.
        cursor.execute(
            "UPDATE categories SET content = ? WHERE id = ? AND wallet_address = ?",
            (content, cat_id, wallet_address)
        )
        conn.commit()
        conn.close()
        return jsonify({"message": "원본 텍스트 수정 완료"}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"DB 에러: {str(e)}"}), 500

# ==========================================
# 💡 텍스트 추출 및 정답(빨간색) 자동 감지 엔진
# ==========================================
def extract_text_from_file(file_obj):
    if not file_obj: return ""
    if file_obj.filename.lower().endswith('.pdf'):
        try:
            doc = fitz.open(stream=file_obj.read(), filetype="pdf")
            text = "".join([page.get_text("text") for page in doc])
            doc.close()
            return text
        except:
            return ""
    else:
        return file_obj.read().decode('utf-8', errors='ignore')

def extract_exam_text_with_color(file_obj, is_answer=False):
    if not file_obj: return ""
    if file_obj.filename.lower().endswith('.pdf'):
        try:
            doc = fitz.open(stream=file_obj.read(), filetype="pdf")
            full_text = ""
            for page in doc:
                dict_data = page.get_text("dict", sort=True)
                for block in dict_data.get("blocks", []):
                    if block.get("type") == 1: 
                        continue
                    for line in block.get("lines", []):
                        line_text = ""
                        for span in line.get("spans", []):
                            text = span.get("text", "").strip()
                            if not text:
                                continue
                            color = span.get("color", 0)
                            r = (color >> 16) & 0xFF
                            g = (color >> 8) & 0xFF
                            b = color & 0xFF
                            if r > 130 and r > g * 1.5 and r > b * 1.5:
                                line_text += f"[🔴{text}]"
                            else:
                                line_text += text
                        if line_text:
                            full_text += line_text + "\n"
                full_text += "\n"
            doc.close()
            return full_text
        except Exception as e:
            print(f"[PDF 추출 오류] {e}", file=sys.stderr)
            return ""
    else:
        return file_obj.read().decode('utf-8', errors='ignore')

def parse_answers_from_text(text: str, question_count: int) -> list:
    ans_dict = {}
    num_map = {'①': '1', '②': '2', '③': '3', '④': '4', '⑤': '5',
               '1': '1', '2': '2', '3': '3', '4': '4', '5': '5'}

    red_matches = re.findall(r'\[🔴([①②③④⑤1-5])\]', text)
    if red_matches:
        for i, ans in enumerate(red_matches):
            ans_dict[i + 1] = num_map.get(ans, ans)

    table_matches = re.findall(r'(\d+)\s*[.)\s]\s*([①②③④⑤1-5])\b', text)
    for num, ans in table_matches:
        n = int(num)
        if 1 <= n <= question_count + 5:
            ans_dict[n] = num_map.get(ans, ans)

    if not ans_dict:
        seq_matches = re.findall(r'[①②③④⑤]', text)
        if len(seq_matches) >= question_count // 2:
            for i, ans in enumerate(seq_matches[:question_count]):
                ans_dict[i + 1] = num_map.get(ans, ans)

    return [ans_dict.get(i + 1, '') for i in range(question_count)]

@api_bp.route('/delete-pending-exam', methods=['POST'])
def delete_pending_exam():
    try:
        data = request.json or {}
        pending_id = data.get('id')
        wallet_address = data.get('wallet_address')
        
        if not pending_id or not wallet_address: return jsonify({"error": "권한 없음"}), 400

        conn = get_db_connection()
        conn.execute("DELETE FROM pending_exams WHERE id = ? AND wallet_address = ?", (pending_id, wallet_address))
        conn.commit()
        conn.close()
        return jsonify({"message": "대기열에서 삭제 완료"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==========================================
# 💡 모의고사 CBT 변환 엔진
# ==========================================
@api_bp.route('/upload-exam-cbt', methods=['POST'])
def upload_exam_cbt():
    import traceback
    import sys
    import requests
    import json
    
    try:
        wallet_address = request.form.get('wallet_address')
        file = request.files.get('file')
        
        if not file or not wallet_address:
            return jsonify({"error": "파일이나 인증 정보가 없습니다."}), 400
            
        exam_text = extract_text_from_file(file)
        if not exam_text:
            return jsonify({"error": "파일에서 텍스트를 추출하지 못했습니다."}), 400
            
        safe_exam_text = exam_text[:2500].replace('"', "'")
        
        prompt = (
            "당신은 아주 정확한 데이터 추출기입니다.\n"
            "아래 [문서 내용]을 읽고, 최대 5개의 객관식 문제를 찾아 JSON 배열로만 출력하세요.\n\n"
            "[문서 내용]\n"
            f"{safe_exam_text}\n\n"
            "[🚨 절대 준수 규칙 🚨]\n"
            "1. JSON 문법을 완벽히 지키세요.\n"
            "2. 해설 내용에 줄바꿈이 필요하면 반드시 '<br>'을 쓰세요. 엔터키는 금지입니다.\n"
            "3. 아래의 마크다운 형식을 그대로 복사해서 내용을 채워 넣는 것으로 출력을 시작하세요.\n\n"
            "```json\n"
            "{\n"
            '  "questions": [\n'
            "    {\n"
            '      "id": 1,\n'
            '      "questionText": "다음 중 ~은?",\n'
            '      "choices": ["보기1", "보기2", "보기3", "보기4"],\n'
            '      "correctAnswer": 0,\n'
            '      "explanation": "해설입니다."\n'
            "    }\n"
            "  ]\n"
            "}\n"
            "```\n"
            "자, 지금 바로 위 형식대로 출력을 시작하세요:"
        )

        print("🤖 [gemma4:26b] 모의고사 파싱 시작 (경량화 모드)...", file=sys.stderr)
        url = "http://localhost:11434/api/generate"
        payload = {
            "model": "gemma4:26b",
            "prompt": prompt,
            "stream": False,
            "options": {
                "temperature": 0.1,
                "num_ctx": 4096,     
                "num_predict": 1500, 
                "top_k": 40,
                "top_p": 0.9,
                "repeat_penalty": 1.15
            }
        }
        
        resp = requests.post(url, json=payload, timeout=600)
        resp.raise_for_status()
        
        response_text = resp.json().get("response", "").strip()
        
        if not response_text:
            raise ValueError("AI가 여전히 빈 응답을 반환했습니다. 모델(gemma4:26b)이 현재 상태에서 응답할 수 없습니다.")

        result = clean_and_parse_json(response_text)
        
        if isinstance(result, dict) and "questions" in result:
            for q in result["questions"]:
                if "explanation" in q and isinstance(q["explanation"], str):
                    q["explanation"] = q["explanation"].replace("<br>", "\n")
                if "questionText" in q and isinstance(q["questionText"], str):
                    q["questionText"] = q["questionText"].replace("<br>", "\n")
                    
        print("✅ [gemma4:26b] 변환 성공!", file=sys.stderr)
        return jsonify(result)
        
    except Exception as e:
        print(f"\n[🔥 CBT 모의고사 변환 에러]\n{traceback.format_exc()}\n", file=sys.stderr)
        return jsonify({"error": str(e)}), 500

@api_bp.route('/upload-exam-coop', methods=['POST'])
def upload_exam_coop():
    try:
        wallet_address = request.form.get('wallet_address')
        exam_file = request.files.get('exam_file')
        answer_file = request.files.get('answer_file')

        if not exam_file or not wallet_address:
            return jsonify({"error": "문제 파일이나 인증 정보가 없습니다."}), 400

        exam_text = extract_exam_text_with_color(exam_file)
        exam_text = re.sub(r'-\s*\d+\s*-', '', exam_text)
        exam_text = re.sub(r'【[^】]+】', '', exam_text)

        exam_text = re.split(r'[\*＊]?\s*단\s*답\s*형', exam_text)[0]

        chunks = re.split(r'(?m)^(?=\s*(?:문\s*)?\d+\s*[.)]\s*[^\s\d])', exam_text)
        valid_chunks = [
            c.strip() for c in chunks
            if c.strip() and len(c.strip()) > 10
            and re.search(r'[①②③④⑤]', c)
        ]

        answers = []
        if answer_file:
            answer_text = extract_exam_text_with_color(answer_file)
            answers = parse_answers_from_text(answer_text, len(valid_chunks))
        else:
            answer_section = exam_text 
            answers = parse_answers_from_text(answer_section, len(valid_chunks))

        conn = get_db_connection()
        conn.execute(
            "INSERT INTO pending_exams (wallet_address, filename, chunks_json, answers_json) VALUES (?, ?, ?, ?)",
            (wallet_address, exam_file.filename,
             json.dumps(valid_chunks, ensure_ascii=False),
             json.dumps(answers, ensure_ascii=False))
        )
        conn.commit()
        conn.close()
        has_answers = any(answers)
        return jsonify({
            "message": "업로드 완료",
            "question_count": len(valid_chunks),
            "answer_count": sum(1 for a in answers if a),
            "has_answers": has_answers
        })
    except Exception as e:
        print(f"[업로드 오류] {traceback.format_exc()}", file=sys.stderr)
        return jsonify({"error": str(e)}), 500

@api_bp.route('/get-pending-exams', methods=['GET'])
def get_pending_exams():
    try:
        wallet_address = request.args.get('wallet_address')
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, filename, chunks_json, answers_json FROM pending_exams WHERE wallet_address = ? ORDER BY id DESC",
            (wallet_address,)
        )
        results = []
        for r in cursor.fetchall():
            results.append({
                "id": r[0],
                "filename": r[1],
                "chunks": json.loads(r[2]),
                "answers": json.loads(r[3]) if r[3] else []
            })
        conn.close()
        return jsonify(results)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@api_bp.route('/cbt-explain', methods=['POST'])
def cbt_explain():
    try:
        data = request.json
        question_text = data.get('question', '')[:1500]
        correct_answer = data.get('correct_answer', '')
        user_answer = data.get('user_answer', '')
        wallet_address = data.get('wallet_address', '')

        char_map = {
            '㉠':'(가)','㉡':'(나)','㉢':'(다)','㉣':'(라)','㉤':'(마)',
            '①':'1번','②':'2번','③':'3번','④':'4번','⑤':'5번',
            '‧':'·','\u200b':'','\xa0':' ',
        }
        for s, d in char_map.items():
            question_text = question_text.replace(s, d)

        db_context = ""
        if wallet_address:
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT folder_name, title, content FROM categories WHERE wallet_address = ?", (wallet_address,))
                rows = cursor.fetchall()
                conn.close()
                if rows:
                    q_words = set(re.findall(r'[가-힣]{2,}', question_text))
                    law_names = set(re.findall(r'[가-힣]+법', question_text))
                    article_nums = set(re.findall(r'제\s*\d+조', question_text))
                    scored = []
                    for folder, title, content in rows:
                        cc = re.sub(r'(?<=[가-힣])\s+(?=[가-힣])', '', content or '')
                        ct = re.sub(r'(?<=[가-힣])\s+(?=[가-힣])', '', title or '')
                        cf = re.sub(r'(?<=[가-힣])\s+(?=[가-힣])', '', folder or '')
                        score = len(q_words & set(re.findall(r'[가-힣]{2,}', cc)))
                        score += len(q_words & set(re.findall(r'[가-힣]{2,}', ct))) * 3
                        for art in article_nums:
                            if re.sub(r'\s','',art) in re.sub(r'\s','',ct): score += 20
                        for law in law_names:
                            if law in cf: score += 30
                        if score > 0:
                            scored.append((score, folder, title, cc[:400]))
                    scored.sort(reverse=True)
                    if scored:
                        db_context = "\n[참고 DB 자료]\n" + "\n".join(
                            f"---\n[{f}/{t}]\n{c}" for _,f,t,c in scored[:5]
                        )
            except Exception as e:
                print(f"[DB 조회 오류] {e}", file=sys.stderr)

        is_wrong = user_answer != correct_answer
        prompt = (
            f"당신은 시험 문제 해설 전문가입니다.\n"
            f"{db_context}\n\n"
            f"[문제]\n{question_text}\n\n"
            f"정답: {correct_answer}번\n"
            f"응시자 답: {user_answer}번 ({'틀림' if is_wrong else '맞음'})\n\n"
            f"정답({correct_answer}번)이 맞는 이유를 DB 자료를 우선 참고하여 3~5문장으로 해설하세요."
        )

        try:
            url = "http://localhost:11434/api/chat"
            payload = {
                "model": "gemma4:26b",
                "messages": [{"role": "user", "content": prompt}],
                "stream": False,
                "think": False,
                "options": {"num_ctx": 4096, "num_predict": 600, "temperature": 0.2}
            }
            resp = requests.post(url, json=payload, timeout=300)
            raw = (resp.json().get("message") or {}).get("content", "").strip()
            raw = re.sub(r'<think>.*?</think>', '', raw, flags=re.DOTALL).strip()
        except Exception as e:
            raw = f"AI 통신 오류: {e}"

        return jsonify({"explanation": raw or "해설 생성 실패"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@api_bp.route('/analyze-chunk', methods=['POST'])
def analyze_chunk():
    try:
        data = request.json
        chunk_text = data.get('chunk_text', '')[:2000]
        user_feedback = data.get('user_feedback', '')
        chat_history = data.get('chat_history', [])

        char_map = {
            '㉠': '(가)', '㉡': '(나)', '㉢': '(다)', '㉣': '(라)',
            '㉤': '(마)', '㉥': '(바)', '㉦': '(사)', '㉧': '(아)',
            '㉨': '(자)', '㉩': '(차)', '㉪': '(카)', '㉫': '(타)',
            '①': '1번', '②': '2번', '③': '3번', '④': '4번', '⑤': '5번',
            '‧': '·', '\u200b': '', '\xa0': ' ', '\u3000': ' ',
            '\u202f': ' ', '\ufeff': '',
        }
        def clean_text(text):
            for src, dst in char_map.items():
                text = text.replace(src, dst)
            text = ''.join(ch if ch == '\n' or ch == '\t' or ch >= ' ' else ' ' for ch in text)
            import re as _re
            text = _re.sub(r'[ \t]{3,}', '  ', text)
            return text.strip()

        chunk_text = clean_text(chunk_text)
        user_feedback = clean_text(user_feedback)

        wallet_address = data.get('wallet_address', '')
        db_context = ""
        if wallet_address:
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT folder_name, title, content FROM categories WHERE wallet_address = ?",
                    (wallet_address,)
                )
                rows = cursor.fetchall()
                conn.close()

                if rows:
                    query_words = set(re.findall(r'[가-힣]{2,}', chunk_text + user_feedback))
                    article_nums = set(re.findall(r'제\s*\d+조', chunk_text + user_feedback))
                    law_names = set(re.findall(r'[가-힣]+법', chunk_text))

                    scored = []
                    for folder, title, content in rows:
                        clean_content = re.sub(r'(?<=[가-힣])\s+(?=[가-힣])', '', content or '')
                        clean_title = re.sub(r'(?<=[가-힣])\s+(?=[가-힣])', '', title or '')
                        clean_folder = re.sub(r'(?<=[가-힣])\s+(?=[가-힣])', '', folder or '')

                        content_words = set(re.findall(r'[가-힣]{2,}', clean_content))
                        title_words = set(re.findall(r'[가-힣]{2,}', clean_title))
                        score = len(query_words & content_words) + len(query_words & title_words) * 3

                        for art in article_nums:
                            art_clean = re.sub(r'\s', '', art)
                            if art_clean in re.sub(r'\s', '', clean_title):
                                score += 20

                        for law in law_names:
                            if law in clean_folder:
                                score += 30

                        if score > 0:
                            scored.append((score, clean_folder, title, clean_content[:600]))

                    scored.sort(reverse=True)
                    top = scored[:10]

                    if top:
                        db_context = "\n\n[참고 DB 자료 - 아래 원문을 최우선으로 참고하세요]\n"
                        for _, folder, title, content in top:
                            db_context += f"---\n[{folder} / {title}]\n{content}\n"
            except Exception as e:
                print(f"[DB 조회 오류] {e}", file=sys.stderr)

        history_lines = []
        for msg in chat_history[-4:]:
            role = "사용자" if msg['sender'] == 'user' else "AI"
            history_lines.append(f"{role}: {msg['text'][:200]}")
        history_str = "\n".join(history_lines)

        history_context = f"\n이전 대화:\n{history_str}\n" if history_str else ""
        user_content = (
            f"당신은 시험 문제 해설 전문가입니다.\n"
            f"아래 [참고 DB 자료]를 최우선으로 활용하여 문제를 분석하세요.\n\n"
            f"[중요 분석 지침]\n"
            f"- 문제에서 ㉠㉡㉢ 등 원문자로 표시된 부분은 '밑줄 친 부분'입니다.\n"
            f"- 질문이 '틀린 것은 몇 개인가'라면, 가~라 단락이 아니라 ㉠~㉩ 각각의 표현이 원문과 맞는지 하나씩 대조하세요.\n"
            f"- 반드시 ㉠부터 ㉩까지 각각 O/X로 판별하고, X인 것의 개수를 세어 최종 정답 번호를 고르세요.\n\n"
            f"{db_context}\n"
            f"[시험 문제]\n{chunk_text}\n"
            f"{history_context}\n"
            f"질문: {user_feedback}"
        )

        try:
            url = "http://localhost:11434/api/chat"
            payload = {
                "model": "gemma4:26b",
                "messages": [{"role": "user", "content": user_content}],
                "stream": False,
                "think": False,
                "options": {
                    "num_ctx": 4096,
                    "num_predict": 1000,
                    "temperature": 0.2,
                    "repeat_penalty": 1.3,
                }
            }
            resp = requests.post(url, json=payload, timeout=300)
            resp.raise_for_status()
            ollama_resp = resp.json()
            raw_text = (ollama_resp.get("message") or {}).get("content", "").strip()
            raw_text = re.sub(r'<think>.*?</think>', '', raw_text, flags=re.DOTALL).strip()
        except Exception as e:
            return jsonify({"error": f"AI 통신 오류: {str(e)}"}), 500

        if not raw_text:
            return jsonify({"error": "AI가 빈 응답을 반환했습니다."}), 500

        if re.search(r'(\b\w+\b)(?:\s+\1){4,}', raw_text):
            return jsonify({"error": "AI가 비정상적인 응답을 생성했습니다. 다시 시도해주세요."}), 500

        answer = "확인 필요"
        ans_match = re.search(r'정답[은이]?\s*[:\：]?\s*([①②③④⑤1-5]번?|확인\s*필요)', raw_text)
        if ans_match:
            answer = ans_match.group(1).strip()

        return jsonify({"result": {
            "chat_message": raw_text,
            "answer": answer,
            "explanation": raw_text
        }})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@api_bp.route('/save-golden-exam', methods=['POST'])
def save_golden_exam():
    try:
        data = request.json
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''INSERT INTO golden_exams 
            (wallet_address, title, question, options_json, answer, explanation, category, search_process, referenced_laws) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (data.get('wallet_address'), data.get('title'), data.get('question'), 
             json.dumps(data.get('options', []), ensure_ascii=False), 
             data.get('answer'), data.get('explanation'), data.get('category', '기본분류'),
             data.get('search_process', ''), data.get('referenced_laws', '')))
        conn.commit()
        conn.close()
        return jsonify({"message": "골든 DB 저장 완료!"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@api_bp.route('/get-golden-exams', methods=['GET'])
def get_golden_exams():
    try:
        wallet_address = request.args.get('wallet_address')
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, title, question, options_json, answer, explanation, search_process, referenced_laws FROM golden_exams WHERE wallet_address = ? ORDER BY id DESC", (wallet_address,))
        
        exams = []
        for r in cursor.fetchall():
            exams.append({
                "id": r[0], "title": r[1], "question": r[2], 
                "options": json.loads(r[3] or "[]"), "answer": r[4], 
                "explanation": r[5], "search_process": r[6], "referenced_laws": r[7]
            })
        conn.close()
        return jsonify({"exams": exams})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@api_bp.route('/get-cbt-session', methods=['GET'])
def get_cbt_session():
    try:
        wallet_address = request.args.get('wallet_address')
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, title, question, options_json, answer, explanation FROM golden_exams WHERE wallet_address = ?", (wallet_address,))
        rows = cursor.fetchall()
        conn.close()
        
        if not rows:
            return jsonify({"error": "골든 DB에 저장된 문제가 없습니다."}), 404
            
        problems = []
        for r in rows:
            problems.append({
                "id": r[0], "title": r[1], "question": r[2], 
                "options": json.loads(r[3] or "[]"), "answer": str(r[4]), "explanation": r[5]
            })
            
        selected = random.sample(problems, min(100, len(problems)))
        return jsonify(selected)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@api_bp.route('/upload-pdf', methods=['POST'])
def upload_pdf():
    file = request.files.get('file')
    wallet_address = request.form.get('wallet_address')
    if not file or not wallet_address:
        return jsonify({"error": "파일 또는 지갑 주소 누락"}), 400

    task_id = str(uuid.uuid4())
    TASK_STATUS[task_id] = "처리 중..."
    
    original_filename = file.filename if file else "일반 규정"
    folder_name = re.sub(r'\.(pdf|txt)$', '', original_filename, flags=re.IGNORECASE)

    def process_file():
        try:
            raw_text = ""
            if original_filename.lower().endswith(('.txt', '.html', '.htm')):
                file_bytes = file.read()
                try: raw_text = file_bytes.decode('utf-8')
                except UnicodeDecodeError: raw_text = file_bytes.decode('cp949', errors='ignore')
            else:
                doc = fitz.open(stream=file.read(), filetype="pdf")
                for page in doc: raw_text += page.get_text()
            
            raw_text = re.sub(r'(?:<|&lt;|〈|＜|\[)\s*(?:신설|개정|삭제|단서신설|전문개정|본조신설|일부개정)[\s\S]*?(?:>|&gt;|〉|＞|\])', '', raw_text)
            raw_text = re.sub(r'(?:\[|［|【)\s*(?:전문개정|본조신설|제목개정|종전제\d+조는|제\d+조에서 이동)[\s\S]*?(?:\]|］|】)', '', raw_text)

            cleaned_text = clean_korean_law_text(raw_text)
            blocks = parse_html_3col_law(cleaned_text)
            
            if not blocks or len(blocks) < 3:
                logging.info(f"[{folder_name}] 일반 문서 파서로 정밀 분석을 시작합니다.")
                blocks = []
                
                pattern = r'(?m)^ *(제\s*\d+\s*조(?:의\s*\d+)?)'
                parts = re.split(pattern, raw_text)
                
                if len(parts) >= 3:
                    for i in range(1, len(parts), 2):
                        article_num = parts[i].strip()
                        content_body = parts[i+1].strip() if i+1 < len(parts) else ""
                        
                        match = re.match(r'^(\s*\(.*?\))', content_body)
                        full_title = f"{article_num} {match.group(1).strip()}" if match else article_num
                        clean_body = re.sub(r'\n{2,}', '\n', content_body).strip()
                        
                        blocks.append({"title": full_title, "content": f"{full_title}\n{clean_body}"})
                else:
                    paragraphs = [p.strip() for p in raw_text.split('\n\n') if len(p.strip()) > 30]
                    for idx, p in enumerate(paragraphs):
                        blocks.append({"title": f"문서 조각 {idx+1}", "content": p})
            
            conn = get_db_connection()
            cursor = conn.cursor()
            for block in blocks:
                block_folder = block.get('folder_name', folder_name)
                cursor.execute(
                    "INSERT INTO categories (wallet_address, title, content, folder_name) VALUES (?, ?, ?, ?)", 
                    (wallet_address, block['title'], block['content'], block_folder)
                )
            conn.commit()
            conn.close()
            TASK_STATUS[task_id] = "완료"
        except Exception as e:
            logging.error(f"분석 에러: {traceback.format_exc()}")
            TASK_STATUS[task_id] = f"에러: {str(e)}"

    threading.Thread(target=process_file).start()
    return jsonify({"message": f"{folder_name} 분석 시작", "task_id": task_id})
    
@api_bp.route('/get-categories')
def get_categories():
    try:
        wallet_address = request.args.get('wallet_address')
        conn = get_db_connection()
        cursor = conn.cursor()
        # 💡 [정렬 기준 수정] sort_order를 최우선 기준으로 정렬합니다.
        cursor.execute("SELECT id, title, content, folder_name FROM categories WHERE wallet_address = ? ORDER BY sort_order ASC, id ASC", (wallet_address,))
        cats = [{"id": r[0], "title": r[1], "content": r[2], "folder_name": r[3]} for r in cursor.fetchall()]
        conn.close()
        return jsonify({"categories": cats})
    except Exception as e:
        return jsonify({"error": "조회 실패"}), 500

@api_bp.route('/split-category', methods=['POST'])
def split_category():
    try:
        data = request.json
        wallet_address = data.get('wallet_address')
        cat_id = data.get('id')
        text1, text2 = data.get('text1'), data.get('text2')
        title1, title2 = data.get('title1'), data.get('title2')
        folder_name = data.get('folder_name') or '기본 폴더'

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM categories WHERE id = ? AND wallet_address = ?", (cat_id, wallet_address))
        cursor.execute("INSERT INTO categories (wallet_address, title, content, folder_name) VALUES (?, ?, ?, ?)", (wallet_address, title1, text1, folder_name))
        cursor.execute("INSERT INTO categories (wallet_address, title, content, folder_name) VALUES (?, ?, ?, ?)", (wallet_address, title2, text2, folder_name))
        conn.commit()
        conn.close()
        return jsonify({"message": "본문 분할 완료"})
    except Exception as e:
        return jsonify({"error": "분할 실패"}), 500

# 💡 [핵심 버그 수정] 라우팅 중복 선언 방지
@api_bp.route('/save-card', methods=['POST'])
def save_card():
    try:
        data = request.json
        wallet_address = data.get('wallet_address')
        card_id = data.get('card_id')
        card_content = data.get('card_content')
        answer_text = data.get('answer_text')
        folder_name = data.get('folder_name', '기본 폴더')
        memo = data.get('memo', '')
        
        if not wallet_address:
            return jsonify({"error": "지갑 주소가 없습니다."}), 400

        conn = get_db_connection()
        cursor = conn.cursor()
        
        if card_id:
            print(f"DEBUG: 카드 수정 시도 - ID: {card_id}, wallet: {wallet_address}")
            cursor.execute('''UPDATE cards SET card_content=?, answer_text=?, folder_name=?, memo=? 
                              WHERE id=? AND wallet_address=?''', 
                              (card_content, answer_text, folder_name, memo, card_id, wallet_address))
            # 💡 수정된 행이 있는지 확인
            print(f"DEBUG: 수정된 행의 수: {cursor.rowcount}")
        else:
            print("DEBUG: 카드 신규 생성 시도")
            cursor.execute('''INSERT INTO cards (wallet_address, category_id, card_content, answer_text, options_json, level, next_review_time, status, best_time, folder_name, memo) 
                              VALUES (?, 0, ?, ?, '[]', 0, ?, 'OWNED', NULL, ?, ?)''', 
                              (wallet_address, card_content, answer_text, get_next_review_time(0), folder_name, memo))
        
        conn.commit()
        conn.close()
        return jsonify({"message": "카드 저장 완료"}), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"DB 에러: {str(e)}"}), 500

@api_bp.route('/my-cards')
def get_my_cards():
    try:
        wallet_address = request.args.get('wallet_address')
        conn = get_db_connection()
        cursor = conn.cursor()
        # 💡 [정렬 기준 수정] sort_order를 최우선 기준으로 정렬합니다.
        cursor.execute("SELECT id, card_content, answer_text, options_json, level, next_review_time, status, best_time, folder_name, memo FROM cards WHERE wallet_address = ? ORDER BY sort_order ASC, id ASC", (wallet_address,))
        cards = [{"id": r[0], "content": r[1], "answer": r[2], "options": json.loads(r[3]), "level": r[4], "next_review_time": r[5], "status": r[6], "best_time": r[7], "folder_name": r[8], "memo": r[9] or ""} for r in cursor.fetchall()]
        conn.close()
        return jsonify({"cards": cards})
    except Exception as e:
        return jsonify({"error": "조회 실패"}), 500
        
@api_bp.route('/delete-category', methods=['POST'])
def delete_category():
    try:
        data = request.json
        wallet_address = data.get('wallet_address')
        cat_id = data.get('id')
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM categories WHERE id = ? AND wallet_address = ?", (cat_id, wallet_address))
        conn.commit()
        conn.close()
        return jsonify({"message": "삭제 완료"})
    except Exception as e:
        return jsonify({"error": "삭제 실패"}), 500

@api_bp.route('/delete-card', methods=['POST'])
def delete_card():
    try:
        data = request.json
        wallet_address = data.get('wallet_address')
        card_id = data.get('id')
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM cards WHERE id = ? AND wallet_address = ?", (card_id, wallet_address))
        conn.commit()
        conn.close()
        return jsonify({"message": "삭제 완료"})
    except Exception as e:
        return jsonify({"error": "삭제 실패"}), 500

@api_bp.route('/delete-all', methods=['POST'])
def delete_all():
    try:
        wallet_address = request.json.get('wallet_address')
        conn = get_db_connection()
        cursor = conn.cursor()
        for table in ['categories', 'cards', 'exams', 'pending_exams', 'golden_exams']:
            cursor.execute(f"DELETE FROM {table} WHERE wallet_address = ?", (wallet_address,))
        conn.commit()
        conn.close()
        return jsonify({"message": "초기화 성공"})
    except Exception as e:
        return jsonify({"error": "초기화 실패"}), 500

@api_bp.route('/sync-batch', methods=['POST'])
def sync_batch():
    try:
        data = request.json
        wallet_address = data.get('wallet_address')
        memos = data.get('memos', [])
        answers = data.get('answers', [])
        
        if not wallet_address: return jsonify({"error": "인증 정보 없음"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        for m in memos:
            cursor.execute("UPDATE cards SET memo = ? WHERE id = ? AND wallet_address = ?", (m.get('memo', ''), m.get('id'), wallet_address))

        reward_coins = 0 

        for a in answers:
            card_id = a.get('card_id')
            is_correct = a.get('is_correct')
            clear_time = float(a.get('clear_time', 999.0))

            cursor.execute("SELECT level, best_time FROM cards WHERE id = ? AND wallet_address = ?", (card_id, wallet_address))
            row = cursor.fetchone()
            # 💡 [추가할 부분] 이 4줄을 바로 아래에 복사해서 붙여넣으세요!
            if not row:
                # 현재 지갑 주소에 데이터가 없다면, DB에 존재하는 다른 주소의 옛날 데이터를 강제로 찾아옵니다.
                cursor.execute("SELECT custom_stopwords, custom_abbrs, custom_inclusions FROM user_settings WHERE custom_stopwords IS NOT NULL AND custom_stopwords != '[]' LIMIT 1")
                row = cursor.fetchone()
                        
            conn.close()
        
            if row:
                current_lv, best_time = row[0], row[1]
                if is_correct:
                    new_lv = min(int(current_lv) + 1, 50)
                    try: best_time_float = float(best_time) if best_time is not None else float('inf')
                    except: best_time_float = float('inf')
                    new_best = clear_time if best_time_float == float('inf') else min(best_time_float, clear_time)
                    cursor.execute("UPDATE cards SET level = ?, next_review_time = ?, status = 'OWNED', best_time = ? WHERE id = ?", (new_lv, get_next_review_time(new_lv), new_best, card_id))
                    reward_coins += 10 
                else:
                    cursor.execute("UPDATE cards SET level = 0, next_review_time = ?, status = 'AT_RISK' WHERE id = ?", (get_next_review_time(0), card_id))

        conn.commit()
        conn.close()

        if reward_coins > 0:
            threading.Thread(target=mint_goal_coin_to_user, args=(wallet_address, reward_coins)).start()

        return jsonify({"message": f"동기화 완료 (지급예정 GOAL: {reward_coins})"}), 200
    except Exception as e:
        return jsonify({"error": "배치 동기화 실패"}), 500

@api_bp.route('/submit-answer', methods=['POST'])
def submit_answer():
    try:
        data = request.json
        wallet_address = data.get('wallet_address')
        card_id = data.get('card_id')
        is_correct = data.get('is_correct')
        try: clear_time = float(data.get('clear_time', 999.0))
        except: clear_time = 999.0

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT level, best_time FROM cards WHERE id = ? AND wallet_address = ?", (card_id, wallet_address))
        row = cursor.fetchone()
        if not row: return jsonify({"error": "카드가 없습니다."}), 404
        current_lv, best_time = row[0], row[1]
        
        msg = ""
        if is_correct:
            new_lv = min(int(current_lv) + 1, 50)
            try: best_time_float = float(best_time) if best_time is not None else float('inf')
            except: best_time_float = float('inf')
            new_best = clear_time if best_time_float == float('inf') else min(best_time_float, clear_time)
            cursor.execute("UPDATE cards SET level = ?, next_review_time = ?, status = 'OWNED', best_time = ? WHERE id = ?", (new_lv, get_next_review_time(new_lv), new_best, card_id))
            msg = f"방어 성공! 레벨이 {new_lv}로 올랐습니다. (10 GOAL 지급 완료)"
            threading.Thread(target=mint_goal_coin_to_user, args=(wallet_address, 10)).start()
        else:
            cursor.execute("UPDATE cards SET level = 0, next_review_time = ?, status = 'AT_RISK' WHERE id = ?", (get_next_review_time(0), card_id))
            msg = "방어 실패! 레벨이 0으로 초기화되었습니다."
            
        conn.commit()
        conn.close()
        return jsonify({"message": msg})
    except Exception as e:
        return jsonify({"error": "제출 처리 실패"}), 500

@api_bp.route('/update-card-memo', methods=['POST'])
def update_card_memo():
    try:
        data = request.json
        wallet_address = data.get('wallet_address')
        card_id = data.get('id')
        memo = data.get('memo', '')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE cards SET memo = ? WHERE id = ? AND wallet_address = ?", (memo, card_id, wallet_address))
        conn.commit()
        conn.close()
        return jsonify({"message": "메모 업데이트 완료"}), 200
    except Exception as e:
        return jsonify({"error": "메모 업데이트 실패"}), 500

# 💡 [핵심 버그 수정] 사용하지 않거나 충돌을 유발하는 낡은 단어장 함수들(get_stopwords, update_stopwords) 완전히 제거 완료

@api_bp.route('/save-checkpoint', methods=['POST'])
def save_checkpoint():
    try:
        data = request.json
        wallet_address = data.get('wallet_address')
        tab = data.get('tab') 
        last_id = data.get('last_id')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute(f"ALTER TABLE user_settings ADD COLUMN last_{tab}_id INTEGER")
        except sqlite3.OperationalError:
            pass 
            
        cursor.execute("SELECT 1 FROM user_settings WHERE wallet_address = ?", (wallet_address,))
        if cursor.fetchone():
            cursor.execute(f"UPDATE user_settings SET last_{tab}_id = ? WHERE wallet_address = ?", (last_id, wallet_address))
        else:
            cursor.execute(f"INSERT INTO user_settings (wallet_address, last_{tab}_id) VALUES (?, ?)", (wallet_address, last_id))
            
        conn.commit()
        conn.close()
        return jsonify({"message": "체크포인트 저장 완료"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@api_bp.route('/get-checkpoint', methods=['GET'])
def get_checkpoint():
    try:
        wallet_address = request.args.get('wallet_address')
        tab = request.args.get('tab')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute(f"SELECT last_{tab}_id FROM user_settings WHERE wallet_address = ?", (wallet_address,))
            row = cursor.fetchone()
            last_id = row[0] if row and row[0] else None
        except sqlite3.OperationalError:
            last_id = None
            
        conn.close()
        return jsonify({"last_id": last_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@api_bp.route('/get-global-dict', methods=['GET'])
def get_global_dict():
    try:
        wallet_address = request.args.get('wallet_address')
        if not wallet_address:
            return jsonify({"stopwords": [], "abbrs": {}, "inclusions": [], "error": "wallet_address 누락"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT custom_stopwords, custom_abbrs, custom_inclusions FROM user_settings WHERE wallet_address = ?", (wallet_address,))
        row = cursor.fetchone()
        conn.close()

        def force_repair_list(val, fallback_key=None):
            """배열 형태 컬럼 복구. fallback_key: 구형 dict에서 꺼낼 키"""
            if not val:
                return []
            try:
                data = json.loads(val)
                if isinstance(data, list):
                    return data
                if isinstance(data, dict):
                    # fallback_key 우선 (예: 'include', 'stop')
                    if fallback_key and fallback_key in data:
                        return data[fallback_key]
                    # 구형 통합 포맷 {"stop":[], "include":[]} 처리
                    if 'stop' in data:
                        return data['stop']
                    if 'stopwords' in data:
                        return data['stopwords']
                return []
            except Exception:
                return [val] if isinstance(val, str) and val else []

        def force_repair_abbr(val):
            if not val:
                return {}
            try:
                result = json.loads(val)
                return result if isinstance(result, dict) else {}
            except Exception:
                return {}

        raw_stopwords  = row[0] if row else None
        raw_abbrs      = row[1] if row else None
        raw_inclusions = row[2] if row else None

        stopwords  = force_repair_list(raw_stopwords,  fallback_key='stop')
        inclusions = force_repair_list(raw_inclusions, fallback_key='include')

        # 💡 구형: custom_stopwords 하나에 {"stop":[], "include":[]} 통합 저장했던 경우
        # inclusions가 비어 있고, 구형 stopwords에 include 키가 있으면 복구
        if not inclusions and raw_stopwords:
            try:
                old = json.loads(raw_stopwords)
                if isinstance(old, dict) and 'include' in old:
                    inclusions = old['include']
            except Exception:
                pass

        return jsonify({
            "stopwords":  stopwords,
            "abbrs":      force_repair_abbr(raw_abbrs),
            "inclusions": inclusions
        })
    except Exception as e:
        logging.error(f"get_global_dict 오류: {e}")
        return jsonify({"error": str(e)}), 500

@api_bp.route('/update-global-dict', methods=['POST'])
def update_global_dict():
    try:
        data = request.json
        wallet_address = data.get('wallet_address')
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT 1 FROM user_settings WHERE wallet_address = ?", (wallet_address,))
        if cursor.fetchone():
            cursor.execute("UPDATE user_settings SET custom_stopwords = ?, custom_abbrs = ?, custom_inclusions = ? WHERE wallet_address = ?", 
                (json.dumps(data.get('stopwords', []), ensure_ascii=False), 
                 json.dumps(data.get('abbrs', {}), ensure_ascii=False),
                 json.dumps(data.get('inclusions', []), ensure_ascii=False),
                 wallet_address))
        else:
            cursor.execute("INSERT INTO user_settings (wallet_address, custom_stopwords, custom_abbrs, custom_inclusions) VALUES (?, ?, ?, ?)", 
                (wallet_address, 
                 json.dumps(data.get('stopwords', []), ensure_ascii=False), 
                 json.dumps(data.get('abbrs', {}), ensure_ascii=False),
                 json.dumps(data.get('inclusions', []), ensure_ascii=False)))
        conn.commit()
        conn.close()
        return jsonify({"message": "전역 사전 DB 업데이트 완료"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

import pandas as pd
import io

# ==========================================
# 💡 [컬럼 예외 완벽 해결] 내 지갑 주소 기준 데이터 전체 엑셀 다운로드 API
# ==========================================
@api_bp.route('/export-excel', methods=['GET'])
def export_excel_final():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wallet_address = request.args.get('wallet_address')
    target = request.args.get('target', 'all')
    
    print(f"\n[엑셀 다운로드 진단 수신] 지갑주소: {wallet_address} | 대상: {target}")
    
    if not wallet_address:
        return jsonify({"error": "인증되지 않은 접근입니다. 지갑 주소가 누락되었습니다."}), 400

    try:
        # 💡 openpyxl 워크북 객체를 메모리에 다이렉트로 직접 생성 (기본 시트 자동 생성됨)
        wb = openpyxl.Workbook()
        
        # openpyxl 기본 생성 시트 가져오기 및 요약 정보 주입
        ws_default = wb.active
        ws_default.title = "안내_및_요약"
        ws_default.views.sheetView[0].showGridLines = True
        
        ws_default["A1"] = "공급망 계정 식별자:"
        ws_default["B1"] = wallet_address
        ws_default["A2"] = "백업 일시:"
        ws_default["B2"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 디자인 서식 지정
        for cell in [ws_default["A1"], ws_default["A2"]]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 1. 만들기 탭 데이터 처리 (categories)
        if target in ['all', 'categories']:
            # 💡 [진단 해결] categories 테이블에는 memo 컬럼이 없으므로 조회 대상에서 제외하여 정합성을 수복합니다.
            cursor.execute(
                "SELECT id, folder_name, title, content FROM categories WHERE wallet_address = ? ORDER BY id ASC", 
                (wallet_address,)
            )
            rows_cat = cursor.fetchall()
            print(f"[만들기 진단] 추출된 행 개수: {len(rows_cat)}개")
            
            ws_cat = wb.create_sheet(title="만들기_카테고리")
            ws_cat.views.sheetView[0].showGridLines = True
            
            headers = ['id', 'folder_name', 'title', 'content']
            ws_cat.append(headers)
            
            for row in rows_cat:
                ws_cat.append([row[0], row[1], row[2], row[3]])
                
        # 2. 채우기 탭 데이터 처리 (cards)
        if target in ['all', 'cards']:
            # 💡 cards 테이블은 기존 기획서 명세대로 memo 컬럼을 포함하여 온전하게 추출합니다.
            cursor.execute(
                "SELECT id, folder_name, card_content, answer_text, memo FROM cards WHERE wallet_address = ? ORDER BY id ASC", 
                (wallet_address,)
            )
            rows_card = cursor.fetchall()
            print(f"[채우기 진단] 추출된 행 개수: {len(rows_card)}개")
            
            ws_card = wb.create_sheet(title="채우기_카드")
            ws_card.views.sheetView[0].showGridLines = True
            
            headers = ['id', 'folder_name', 'content', 'answer_text', 'memo']
            ws_card.append(headers)
            
            for row in rows_card:
                ws_card.append([row[0], row[1], row[2], row[3], row[4]])
                
        conn.close()
        
        # 가상 바이너리 파일로 스트리밍 빌드
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        print("✅ [성공] 컬럼 매핑 에러 제어 및 가상 통합 문서 빌드 통과 완료.")
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"blankd_내자료_백업_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
    except Exception as e:
        err_msg = traceback.format_exc()
        print(f"❌ [백엔드 치명적 에러 캐치]\n{err_msg}", file=sys.stderr)
        return jsonify({"error": f"서버 내부 연산 장애: {str(e)}", "traceback": err_msg}), 500

# ==========================================
# 💡 [신규 추가] 드래그/버튼 기반 순서 변경 업데이트 API
# ==========================================
@api_bp.route('/update-order', methods=['POST'])
def update_order():
    try:
        data = request.json
        wallet_address = data.get('wallet_address')
        table = data.get('table') 
        ordered_ids = data.get('ordered_ids', [])
        
        if not wallet_address or table not in ['categories', 'cards'] or not ordered_ids:
            return jsonify({"error": "잘못된 요청입니다."}), 400
            
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 전달받은 ID 배열의 인덱스(0, 1, 2...)를 그대로 sort_order에 덮어씌웁니다.
        for index, item_id in enumerate(ordered_ids):
            cursor.execute(f"UPDATE {table} SET sort_order = ? WHERE id = ? AND wallet_address = ?", (index, item_id, wallet_address))
            
        conn.commit()
        conn.close()
        return jsonify({"message": "순서 변경 완료"}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ==========================================
# 💡 [신규 추가] 앱 전체 사용자 빈칸 채우기 랭킹 API
# ==========================================
@api_bp.route('/ranking', methods=['GET'])
def get_ranking():
    import json
    import re
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        # 모든 카드의 소유자(wallet_address)와 memo(학습 통계)를 가져옵니다.
        cursor.execute("SELECT wallet_address, memo FROM cards WHERE wallet_address IS NOT NULL AND wallet_address != ''")
        rows = cursor.fetchall()
        conn.close()

        user_scores = {}
        for row in rows:
            wallet = row[0]
            memo = row[1] or ""
            filled = 0
            
            # 카드별 학습 횟수 추출 (안전한 다중 파싱 지원)
            try:
                parsed = json.loads(memo)
                if isinstance(parsed, dict):
                    filled = int(parsed.get('filled', 0))
            except:
                match = re.search(r'"filled"\s*:\s*(\d+)', memo)
                if match:
                    filled = int(match.group(1))
                else:
                    parts = memo.split('||')
                    if len(parts) >= 2 and parts[1].isdigit():
                        filled = int(parts[1])

            if wallet not in user_scores:
                user_scores[wallet] = 0
            user_scores[wallet] += filled

        # 랭킹 정렬 및 0회 유저 제외 (상위 50명까지만 반환)
        sorted_ranking = sorted(user_scores.items(), key=lambda x: x[1], reverse=True)
        top_ranking = [{"wallet_address": k, "total_filled": v} for k, v in sorted_ranking if v > 0][:50]

        return jsonify({"ranking": top_ranking}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": "랭킹 조회 실패", "details": str(e)}), 500

@api_bp.route('/update-balance', methods=['POST', 'OPTIONS'])
def update_balance():
    if request.method == 'OPTIONS': return jsonify({}), 200
    try:
        data = request.json
        wallet_address = data.get('wallet_address')
        if not wallet_address: return jsonify({"error": "No wallet"}), 400
        
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # 💡 [안전장치] user_settings 테이블에 컬럼이 없을 경우를 대비한 자동 생성
        try: cursor.execute("ALTER TABLE user_settings ADD COLUMN goal_balance INTEGER DEFAULT 0")
        except: pass
        try: cursor.execute("ALTER TABLE user_settings ADD COLUMN activity_log TEXT DEFAULT '{}'")
        except: pass
        try: cursor.execute("ALTER TABLE user_settings ADD COLUMN claimed_rewards TEXT DEFAULT '{}'")
        except: pass
        
        cursor.execute("SELECT goal_balance, activity_log, claimed_rewards FROM user_settings WHERE wallet_address = ?", (wallet_address,))
        row = cursor.fetchone()
        
        if row:
            db_bal = row['goal_balance'] if row['goal_balance'] is not None else 0
            incoming_bal = data.get('balance')
            
            updates, params = [], []
            
            # 💡 [핵심 방어 로직] 들어온 포인트가 기존 서버 포인트보다 "클 때만" 덮어쓰기 허용
            if incoming_bal is not None and incoming_bal > db_bal:
                updates.append("goal_balance = ?")
                params.append(incoming_bal)
                
            if data.get('activity_log') is not None: 
                updates.append("activity_log = ?")
                params.append(json.dumps(data.get('activity_log')))
                
            if data.get('claimed_rewards') is not None: 
                updates.append("claimed_rewards = ?")
                params.append(json.dumps(data.get('claimed_rewards')))
                
            if updates:
                params.append(wallet_address)
                cursor.execute(f"UPDATE user_settings SET {', '.join(updates)} WHERE wallet_address = ?", params)
        else:
            cursor.execute("INSERT INTO user_settings (wallet_address, goal_balance, activity_log, claimed_rewards) VALUES (?, ?, ?, ?)", 
                (wallet_address, data.get('balance') or 0, json.dumps(data.get('activity_log') or {}), json.dumps(data.get('claimed_rewards') or {})))
            
        conn.commit()
        conn.close()
        return jsonify({"message": "완료"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@api_bp.route('/get-balance', methods=['GET', 'OPTIONS'])
def get_balance():
    if request.method == 'OPTIONS': return jsonify({}), 200
    try:
        wallet_address = request.args.get('wallet_address')
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT goal_balance, activity_log, claimed_rewards FROM user_settings WHERE wallet_address = ?", (wallet_address,))
            row = cursor.fetchone()
            balance = row['goal_balance'] if row and row['goal_balance'] is not None else 0
            activity_log = json.loads(row['activity_log']) if row and row['activity_log'] else {}
            claimed_rewards = json.loads(row['claimed_rewards']) if row and row['claimed_rewards'] else {}
        except:
            balance, activity_log, claimed_rewards = 0, {}, {}
        conn.close()
        return jsonify({"balance": balance, "activity_log": activity_log, "claimed_rewards": claimed_rewards}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
